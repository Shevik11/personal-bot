"""Import the consolidated expense sheet from the legacy workbook.

Usage:
    uv run python scripts/import_expenses.py data/Витрати.xlsx --user-id 123 --apply
"""

from __future__ import annotations

import argparse
import asyncio
from datetime import date, datetime
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from telegram_bot import storage


def _parse_date(value: object, epoch: datetime) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        converted = from_excel(value, epoch)
        if isinstance(converted, datetime):
            return converted.date().isoformat()
        if isinstance(converted, date):
            return converted.isoformat()
    return None


def _parse_amount(value: object) -> int | None:
    if value is None:
        return None
    try:
        amount = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return None
    if amount <= 0:
        return None
    return int(amount * 100)


async def import_workbook(path: Path, user_id: int, apply: bool) -> tuple[int, int, int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    imported = skipped = duplicates = 0

    await storage.initialize_database()
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=2, max_col=4, values_only=True), start=2
    ):
        expense_date = _parse_date(row[0], workbook.epoch)
        amount_minor = _parse_amount(row[1])
        merchant = str(row[2]).strip() if row[2] is not None else ""
        description = str(row[3]).strip() if row[3] is not None else ""
        if not expense_date or amount_minor is None or not merchant or not description:
            skipped += 1
            continue

        if await storage.expense_exists(
            user_id,
            expense_date,
            amount_minor,
            merchant,
            description,
        ):
            duplicates += 1
            continue

        if apply:
            await storage.add_expense(
                user_id,
                expense_date,
                amount_minor,
                merchant,
                description,
            )
        imported += 1

    workbook.close()
    return imported, duplicates, skipped


async def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--user-id", type=int, required=True)
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Actually write records; without this flag the import is a dry run.",
    )
    args = parser.parse_args()
    imported, duplicates, skipped = await import_workbook(
        args.workbook, args.user_id, args.apply
    )
    mode = "imported" if args.apply else "would import"
    print(f"{mode} {imported} rows; skipped {duplicates} duplicates and {skipped} invalid rows")


if __name__ == "__main__":
    asyncio.run(main())
